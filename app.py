from flask import Flask, render_template, request, send_file, redirect, url_for, session, flash
import pandas as pd
import os
import io
import json
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from utils.procesamiento import (
    procesar_registros, HORARIOS_SEDES,
    agregar_empleado, cargar_empleados, guardar_empleados
)
from database import (
    init_db, db_get_all, db_get_by_id, db_get_by_nombre,
    db_crear, db_upsert, db_eliminar, db_actualizar,
    db_get_horario, db_guardar_horario, resolver_horario, DIAS_ORDEN,
    db_guardar_permiso, db_obtener_permisos, db_obtener_todos_permisos, db_eliminar_permiso
)

app = Flask(__name__)
app.secret_key = "clave-ultra-secreta"

CONFIG_PASSWORD = "admin123"
MEMORY = {"permisos": []}
UPLOAD_FOLDER = "uploads"
SETTINGS_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "config", "settings.json"))
SETTINGS = None
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Inicializar BD al arrancar
init_db()


def cargar_settings():
    global SETTINGS
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, encoding="utf-8") as f:
                SETTINGS = json.load(f)
        except Exception:
            SETTINGS = None
    if SETTINGS is None:
        SETTINGS = {"umbral_extras_minutos": 30}
        guardar_settings(SETTINGS)
    return SETTINGS


def guardar_settings(settings_dict):
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(settings_dict, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


SETTINGS = cargar_settings()


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def parse_hhmm(valor):
    if isinstance(valor, str) and "h" in valor:
        partes = valor.replace("h", "").replace("m", "").split()
        try:
            return int(partes[0]) * 60 + int(partes[1])
        except Exception:
            return 0
    return 0

def attach_permissions(df):
    permisos = MEMORY.get("permisos", [])
    if df is None or df.empty or not permisos:
        return df
    df = df.copy()
    df["Permiso"] = ""
    for permiso in permisos:
        nombre_perm = permiso["nombre"].strip().lower()
        fecha_perm  = permiso["fecha"]
        texto_perm  = f"{permiso['tipo']} ({permiso['minutos']}m)"
        if permiso.get("motivo"):
            texto_perm += f" - {permiso['motivo']}"
        for idx, row in df.iterrows():
            if str(row["Nombre"]).strip().lower() == nombre_perm and row["Fecha"] == fecha_perm:
                existente = df.at[idx, "Permiso"]
                df.at[idx, "Permiso"] = f"{existente}; {texto_perm}" if existente else texto_perm
    return df

def agregar_permisos_bd(df):
    """Agrega permisos desde la BD"""
    if df is None or df.empty:
        return df
    df = df.copy()
    df["Permiso"] = ""
    for idx, row in df.iterrows():
        nombre = row.get("Nombre", "")
        fecha = row.get("Fecha", "")
        if not nombre or str(nombre).startswith("TOTAL"):
            continue
        permisos_dict = db_obtener_permisos(nombre, fecha)
        if permisos_dict:
            permiso_strs = []
            for tipo, motivo in permisos_dict.items():
                if motivo:
                    permiso_strs.append(f"{tipo}: {motivo}")
                else:
                    permiso_strs.append(tipo)
            df.at[idx, "Permiso"] = "; ".join(permiso_strs)
    return df

def require_config_auth(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("config_auth"):
            return redirect(url_for("config_login"))
        return f(*args, **kwargs)
    return decorated

def guardar_config_sedes(config_dict):
    ruta = os.path.abspath(os.path.join(os.path.dirname(__file__), "config", "sedes.json"))
    try:
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(config_dict, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def construir_almuerzo_excel(df):
    """Construye un Excel con tablas de almuerzo en grilla (máx 3 por fila) con estilos."""
    if df is None or df.empty:
        return io.BytesIO()

    df = df[~df["Nombre"].str.contains("TOTAL HORAS EXTRAS", na=False)].copy()
    # Filtrar sábados (sin almuerzo)
    df = df[df["Día"] != "Sábado"].copy()
    
    columnas = ["Nombre", "Fecha", "Día", "Sal. Almuerzo", "Reg. Almuerzo", "T. Almuerzo"]
    df = df[[c for c in columnas if c in df.columns]]

    df = df.copy()
    df["Minutos almuerzo"] = 0
    for idx, row in df.iterrows():
        valor = row.get("T. Almuerzo", "")
        if isinstance(valor, str) and "h" in valor:
            partes = valor.replace("h", "").replace("m", "").split()
            try:
                df.at[idx, "Minutos almuerzo"] = int(partes[0]) * 60 + int(partes[1])
            except Exception:
                df.at[idx, "Minutos almuerzo"] = 0

    df = df.sort_values(["Nombre", "Fecha"], kind="mergesort", na_position="last")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        hoja = writer.book.create_sheet("Almuerzo", 0)
        
        empleados = sorted(df["Nombre"].unique())
        max_cols_por_fila = 3
        cols_por_tabla = 6  # 5 columnas de datos + 1 de separación
        
        # Estilos
        estilo_encabezado = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        estilo_fuente_encabezado = Font(bold=True, color="FFFFFF")
        estilo_borde_cell = Alignment(wrap_text=True, vertical="center")
        
        borde_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        fila_actual = 1
        
        # Procesar empleados en grupos de 3
        for grupo_idx in range(0, len(empleados), max_cols_por_fila):
            grupo_empleados = empleados[grupo_idx:grupo_idx + max_cols_por_fila]
            
            # Calcular altura máxima del grupo
            altura_max = 0
            for emp in grupo_empleados:
                grp = df[df["Nombre"] == emp]
                altura = len(grp) + 2  # datos + encabezado + títulos (sin total)
                altura_max = max(altura_max, altura)
            
            # Dibujar tablas lado a lado
            for col_pos, empleado in enumerate(grupo_empleados):
                grp = df[df["Nombre"] == empleado].copy()
                grp_show = grp[["Fecha", "Día", "Sal. Almuerzo", "Reg. Almuerzo", "T. Almuerzo"]]
                
                col_inicio = 1 + (col_pos * cols_por_tabla)
                
                # Encabezado con nombre del empleado
                celda_titulo = hoja.cell(row=fila_actual, column=col_inicio, value=empleado.upper())
                celda_titulo.fill = estilo_encabezado
                celda_titulo.font = estilo_fuente_encabezado
                celda_titulo.border = borde_thin
                
                # Títulos de columnas
                cols_titulo = ["Fecha", "Día", "Sal. Almuerzo", "Reg. Almuerzo", "T. Almuerzo"]
                for i, titulo in enumerate(cols_titulo):
                    celda = hoja.cell(row=fila_actual + 1, column=col_inicio + i, value=titulo)
                    celda.fill = estilo_encabezado
                    celda.font = estilo_fuente_encabezado
                    celda.border = borde_thin
                    celda.alignment = estilo_borde_cell
                
                # Datos
                for fila_idx, (_, fila) in enumerate(grp_show.iterrows(), start=2):
                    for col_idx_offset, col_nombre in enumerate(cols_titulo):
                        valor = fila.get(col_nombre, "")
                        celda = hoja.cell(row=fila_actual + fila_idx, column=col_inicio + col_idx_offset, value=valor)
                        celda.border = borde_thin
                        celda.alignment = estilo_borde_cell
                
                # Ajustar ancho de columnas
                for i in range(5):
                    col_letra = get_column_letter(col_inicio + i)
                    hoja.column_dimensions[col_letra].width = 18
            
            # Pasar a siguiente grupo de filas (altura_max + 2 para separación)
            fila_actual += altura_max + 2

    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────
# RUTAS PRINCIPALES
# ─────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html", sedes=list(HORARIOS_SEDES.keys()))

@app.route("/horarios")
def horarios():
    return render_template("horarios.html", horarios=HORARIOS_SEDES)

@app.route("/subir", methods=["POST"])
def subir():
    archivo = request.files.get("archivo_csv")
    sede    = request.form.get("sede")

    if not archivo or not archivo.filename:
        return render_template("index.html", mensaje_error="No se seleccionó archivo",
                               sedes=list(HORARIOS_SEDES.keys()))
    if sede not in HORARIOS_SEDES:
        return render_template("index.html", mensaje_error="Seleccione una sede válida",
                               sedes=list(HORARIOS_SEDES.keys()))

    ruta = os.path.join(UPLOAD_FOLDER, archivo.filename)
    archivo.save(ruta)
    extension = archivo.filename.lower().split(".")[-1]

    try:
        if extension in ["xlsx", "xls"]:
            df = pd.read_excel(ruta)
        elif extension == "csv":
            try:
                df = pd.read_csv(ruta, encoding="utf-8", sep=None, engine="python", on_bad_lines="skip")
            except:
                df = pd.read_csv(ruta, encoding="latin-1", sep=None, engine="python", on_bad_lines="skip")
        else:
            return render_template("index.html",
                                   mensaje_error="Formato no soportado. Use CSV o Excel.",
                                   sedes=list(HORARIOS_SEDES.keys()))
    except Exception as e:
        return render_template("index.html",
                               mensaje_error=f"Error leyendo el archivo: {str(e)}",
                               sedes=list(HORARIOS_SEDES.keys()))

    # Función que resuelve el horario efectivo por empleado (BD > sede)
    def fn_resolver(nombre, sede_nombre):
        return resolver_horario(nombre, sede_nombre, HORARIOS_SEDES)

    procesado = procesar_registros(
        df, sede, resolver_horario_fn=fn_resolver,
        umbral_extras_minutos=SETTINGS.get("umbral_extras_minutos", 30)
    )
    procesado = attach_permissions(procesado)

    # Importar empleados a la BD si no existen
    for nombre in procesado["Nombre"].unique():
        if "TOTAL" not in str(nombre).upper():
            agregar_empleado(nombre)          # JSON (compatibilidad)
            db_upsert(nombre, sede)           # SQLite

    MEMORY["df"]   = procesado
    MEMORY["sede"] = sede
    return redirect(url_for("vista_previa"))


@app.route("/vista_previa", methods=["GET", "POST"])
def vista_previa():
    df = MEMORY.get("df")
    if df is None:
        return "No hay datos cargados"

    nombres = sorted([n for n in df["Nombre"].unique() if "TOTAL" not in n.upper()])

    if request.method == "GET":
        return render_template("vista_previa.html", tabla=df.to_dict(orient="records"),
                               nombres=nombres, empleado_seleccionado=None)

    empleado = request.form.get("empleado")
    if not empleado:
        return render_template("vista_previa.html", tabla=df.to_dict(orient="records"),
                               nombres=nombres, empleado_seleccionado="")

    filtrado      = df[df["Nombre"] == empleado].copy()
    filas_normales = filtrado[~filtrado["Nombre"].str.contains("TOTAL")].copy()
    total_min     = filas_normales["Horas extras"].apply(parse_hhmm).sum()
    total_str     = f"{total_min // 60:02d}h {total_min % 60:02d}m"

    fila_total = {
        "Nombre": f"TOTAL HORAS EXTRAS ({empleado})",
        "Fecha": "", "Día": "", "Entrada": "",
        "Sal. Almuerzo": "", "Reg. Almuerzo": "", "Salida": "",
        "T. Almuerzo": "", "Horas trabajadas": "", "Tardanza": "TOTAL",
        "Horas extras": total_str, "Validación": "", "Detalle validación": "",
    }
    return render_template("vista_previa.html",
                           tabla=filas_normales.to_dict("records") + [fila_total],
                           nombres=nombres, empleado_seleccionado=empleado)


# ─────────────────────────────────────────────────────────────
# GUARDAR PERMISO (desde vista previa)
# ─────────────────────────────────────────────────────────────

@app.route("/guardar_permiso", methods=["POST"])
def guardar_permiso():
    try:
        data = request.get_json()
        nombre = data.get("nombre", "").strip()
        fecha = data.get("fecha", "").strip()  # DD/MM/YYYY
        motivo = data.get("motivo", "").strip()
        tipo = data.get("tipo", "").strip()  # tardanza o salida_temprana
        
        if not all([nombre, fecha, tipo]):
            return {"success": False, "error": "Datos incompletos"}, 400
        
        # Convertir fecha a formato DD/MM/YYYY si es necesario
        db_guardar_permiso(nombre, fecha, motivo, tipo)
        return {"success": True, "message": "Permiso guardado"}
    except Exception as e:
        return {"success": False, "error": str(e)}, 500


# ─────────────────────────────────────────────────────────────
# PERMISOS
# ─────────────────────────────────────────────────────────────

@app.route("/permisos", methods=["GET", "POST"])
def permisos():
    permisos_list = MEMORY.setdefault("permisos", [])
    mensaje = mensaje_error = None

    if request.method == "POST":
        nombre  = request.form.get("nombre", "").strip()
        fecha   = request.form.get("fecha", "").strip()
        tipo    = request.form.get("tipo", "").strip()
        minutos = request.form.get("minutos", "").strip()
        motivo  = request.form.get("motivo", "").strip()

        if not all([nombre, fecha, tipo, minutos]):
            mensaje_error = "Complete todos los campos obligatorios."
        else:
            try:
                permisos_list.append({
                    "nombre": nombre, "fecha": fecha,
                    "tipo": tipo, "minutos": int(minutos), "motivo": motivo,
                })
                MEMORY["df"] = attach_permissions(MEMORY.get("df"))
                mensaje = "Permiso guardado correctamente."
            except ValueError:
                mensaje_error = "Los minutos deben ser un número entero."

    return render_template("permisos.html", permisos=permisos_list,
                           mensaje=mensaje, mensaje_error=mensaje_error,
                           sedes=list(HORARIOS_SEDES.keys()))


# ─────────────────────────────────────────────────────────────
# EMPLEADOS (BD SQLite)
# ─────────────────────────────────────────────────────────────

@app.route("/empleados")
def empleados():
    return render_template("empleados.html",
                           empleados=db_get_all(),
                           sedes=list(HORARIOS_SEDES.keys()))

@app.route("/empleados/crear", methods=["POST"])
def empleados_crear():
    nombre = request.form.get("nombre", "").strip()
    sede   = request.form.get("sede", "")
    if not nombre or sede not in HORARIOS_SEDES:
        flash("Nombre y sede válidos son obligatorios.", "error")
    else:
        resultado = db_crear(nombre, sede)
        if resultado == -1:
            flash(f'El empleado "{nombre}" ya existe.', "error")
        else:
            agregar_empleado(nombre)  # sincronizar JSON
            flash(f'Empleado "{nombre}" creado.', "success")
    return redirect(url_for("empleados"))

@app.route("/empleados/eliminar/<int:eid>", methods=["POST"])
def empleados_eliminar(eid):
    emp = db_get_by_id(eid)
    if emp:
        db_eliminar(eid)
        flash(f'Empleado "{emp["nombre"]}" eliminado.', "success")
    return redirect(url_for("empleados"))

@app.route("/empleados/editar/<int:eid>", methods=["GET", "POST"])
def empleados_editar(eid):
    emp = db_get_by_id(eid)
    if not emp:
        return redirect(url_for("empleados"))

    horario_actual = db_get_horario(eid)

    if request.method == "POST":
        sede         = request.form.get("sede", emp["sede"])
        usar_custom  = request.form.get("horario_custom") == "1"
        db_actualizar(eid, sede, int(usar_custom))

        if usar_custom:
            nuevo = {}
            for dia in DIAS_ORDEN:
                entrada = request.form.get(f"entrada_{dia}", "").strip()
                salida  = request.form.get(f"salida_{dia}", "").strip()
                if entrada and salida:
                    nuevo[dia] = {"entrada": entrada, "salida": salida}
            db_guardar_horario(eid, nuevo)
        else:
            db_guardar_horario(eid, {})  # limpiar horario custom

        flash("Cambios guardados.", "success")
        return redirect(url_for("empleados"))

    return render_template("empleado_editar.html",
                           emp=emp,
                           dias=DIAS_ORDEN,
                           sedes=list(HORARIOS_SEDES.keys()),
                           horario_actual=horario_actual,
                           sede_horario=HORARIOS_SEDES.get(emp["sede"], {}))


# ─────────────────────────────────────────────────────────────
# TABLA RESULTADOS / DESCARGA RESUMEN
# ─────────────────────────────────────────────────────────────

@app.route("/tabla_resultados")
def tabla_resultados():
    df = MEMORY.get("df")
    if df is None or df.empty:
        return render_template("tabla_resultados.html", columnas=[], filas=[],
                               mensaje="No hay datos cargados.",
                               sedes=list(HORARIOS_SEDES.keys()))

    df_r = df[~df["Nombre"].str.contains("TOTAL HORAS EXTRAS", na=False)].copy()
    filas = []
    for nombre, grp in df_r.groupby("Nombre", sort=False):
        total_trab   = sum(parse_hhmm(v) for v in grp["Horas trabajadas"])
        total_extras = sum(parse_hhmm(v) for v in grp["Horas extras"])
        permisos_text = "; ".join(
            f"{p['tipo']} {p['minutos']}m"
            for p in MEMORY.get("permisos", [])
            if p["nombre"].strip().lower() == nombre.strip().lower()
        )
        validaciones = grp["Validación"].dropna().tolist()
        estado_general = "OK"
        if any(v == "Revisar" for v in validaciones):
            estado_general = "Revisar"
        filas.append({
            "Nombre": nombre,
            "Días registrados": len(grp),
            "Horas trabajadas totales": f"{total_trab // 60:02d}h {total_trab % 60:02d}m",
            "Horas extras totales": f"{total_extras // 60:02d}h {total_extras % 60:02d}m",
            "Validación": estado_general,
            "Permisos": permisos_text,
        })

    columnas = ["Nombre", "Días registrados", "Horas trabajadas totales",
                "Horas extras totales", "Validación", "Permisos"]
    return render_template("tabla_resultados.html", columnas=columnas, filas=filas,
                           mensaje=None, sedes=list(HORARIOS_SEDES.keys()))

@app.route("/descargar_resumen")
def descargar_resumen():
    df = MEMORY.get("df")
    if df is None or df.empty:
        return "No hay datos cargados para descargar", 400

    df_r   = df[~df["Nombre"].str.contains("TOTAL HORAS EXTRAS", na=False)].copy()
    resumen = []
    for nombre, grp in df_r.groupby("Nombre", sort=False):
        total_trab   = sum(parse_hhmm(v) for v in grp["Horas trabajadas"])
        total_extras = sum(parse_hhmm(v) for v in grp["Horas extras"])
        resumen.append({
            "Nombre": nombre,
            "Días registrados": len(grp),
            "Horas trabajadas totales": f"{total_trab // 60:02d}h {total_trab % 60:02d}m",
            "Horas extras totales": f"{total_extras // 60:02d}h {total_extras % 60:02d}m",
        })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(resumen).to_excel(writer, index=False, sheet_name="Resumen")
        hoja     = writer.sheets["Resumen"]
        last_row = len(resumen) + 1
        last_col = len(resumen[0].keys()) if resumen else 1
        tabla    = Table(displayName="TablaResumen", ref=f"A1:{get_column_letter(last_col)}{last_row}")
        tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False)
        hoja.add_table(tabla)
        for col in hoja.columns:
            hoja.column_dimensions[col[0].column_letter].width = \
                max((len(str(c.value or "")) for c in col), default=0) + 3
    output.seek(0)
    return send_file(output, download_name="resumen_horas.xlsx", as_attachment=True)


# ─────────────────────────────────────────────────────────────
# DESCARGAS EXTRAS / LLEGADAS
# ─────────────────────────────────────────────────────────────

@app.route("/descargar_extras")
def descargar_extras():
    nombre = request.args.get("nombre", "todos")
    df     = MEMORY.get("df")
    if df is None or df.empty:
        return "No hay datos cargados", 400

    df = df[~df["Nombre"].str.contains("TOTAL HORAS EXTRAS", na=False)].copy()
    if nombre != "todos":
        df = df[df["Nombre"] == nombre].copy()

    columnas = ["Nombre", "Fecha", "Día", "Entrada", "Sal. Almuerzo",
                "Reg. Almuerzo", "Salida", "T. Almuerzo",
                "Horas trabajadas", "Tardanza", "Horas extras", "Permiso"]
    
    # Agregar permisos desde BD
    df = agregar_permisos_bd(df)
    df = df[[c for c in columnas if c in df.columns]]

    df["extra_min"] = df["Horas extras"].apply(parse_hhmm)
    resumen = []

    if nombre == "todos":
        bloques = []
        for empleado, grp in df.groupby("Nombre", sort=False):
            bloques.append(grp)
            total_min_emp = grp["extra_min"].sum()
            subtotal = {c: "" for c in columnas}
            subtotal.update({
                "Nombre": f"TOTAL HORAS EXTRAS ({empleado})",
                "Tardanza": "TOTAL",
                "Horas extras": f"{total_min_emp // 60:02d}h {total_min_emp % 60:02d}m",
                "extra_min": total_min_emp,
            })
            bloques.append(pd.DataFrame([subtotal]))
            resumen.append({
                "Nombre": empleado,
                "Horas extras totales": f"{total_min_emp // 60:02d}h {total_min_emp % 60:02d}m",
                "extra_min": total_min_emp,
            })
        df_total = pd.concat(bloques, ignore_index=True) if bloques else pd.DataFrame(columns=columnas + ["extra_min"])
        total_min = df["extra_min"].sum()
        fila_total = {c: "" for c in columnas}
        fila_total.update({
            "Nombre": "TOTAL GENERAL",
            "Tardanza": "TOTAL",
            "Horas extras": f"{total_min // 60:02d}h {total_min % 60:02d}m",
            "extra_min": total_min,
        })
        df_total = pd.concat([df_total, pd.DataFrame([fila_total])], ignore_index=True)
        resumen.append({
            "Nombre": "TOTAL GENERAL",
            "Horas extras totales": f"{total_min // 60:02d}h {total_min % 60:02d}m",
            "extra_min": total_min,
        })
    else:
        total_min = df["extra_min"].sum()
        fila_total = {c: "" for c in columnas}
        fila_total.update({
            "Tardanza": "TOTAL",
            "Horas extras": f"{total_min // 60:02d}h {total_min % 60:02d}m",
            "extra_min": total_min,
        })
        df_total = pd.concat([df, pd.DataFrame([fila_total])], ignore_index=True)
        resumen.append({
            "Nombre": nombre,
            "Horas extras totales": f"{total_min // 60:02d}h {total_min % 60:02d}m",
            "extra_min": total_min,
        })

    df_total = df_total.drop(columns=["extra_min"])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Si descarga todos los empleados, crear una hoja por cada uno
        if nombre == "todos":
            # Crear hojas individuales por empleado
            for empleado, grp in df.groupby("Nombre", sort=False):
                # Agregar fila de total para este empleado
                total_min_emp = grp["extra_min"].sum()
                fila_total_emp = {c: "" for c in columnas}
                fila_total_emp.update({
                    "Nombre": f"TOTAL HORAS EXTRAS ({empleado})",
                    "Tardanza": "TOTAL",
                    "Horas extras": f"{total_min_emp // 60:02d}h {total_min_emp % 60:02d}m",
                })
                df_empleado = pd.concat([grp, pd.DataFrame([fila_total_emp])], ignore_index=True)
                df_empleado = df_empleado.drop(columns=["extra_min"])
                
                # Escribir en hoja del empleado (máx 31 caracteres para nombre de hoja)
                nombre_hoja = empleado[:31]
                df_empleado.to_excel(writer, index=False, sheet_name=nombre_hoja)
                
                # Formatear hoja del empleado
                hoja = writer.sheets[nombre_hoja]
                last_row = len(df_empleado) + 1
                last_col = len(df_empleado.columns)
                tabla = Table(displayName=f"Tabla{nombre_hoja.replace(' ', '')}", ref=f"A1:{get_column_letter(last_col)}{last_row}")
                tabla.tableStyleInfo = TableStyleInfo(name="TableStyleLight11",
                    showFirstColumn=False, showLastColumn=False,
                    showRowStripes=False, showColumnStripes=False)
                hoja.add_table(tabla)
                for col in hoja.columns:
                    hoja.column_dimensions[col[0].column_letter].width = \
                        max((len(str(c.value or "")) for c in col), default=0) + 3
                for i in range(1, last_row + 1):
                    hoja.row_dimensions[i].height = 22
                fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                for cell in hoja[last_row]:
                    cell.fill = fill; cell.font = Font(bold=True)
            
            # Agregar hoja de resumen
            resumen_df = pd.DataFrame(resumen).drop(columns=["extra_min"])
            resumen_df.to_excel(writer, index=False, sheet_name="Resumen")
            
            hoja_resumen = writer.sheets["Resumen"]
            last_row_res = len(resumen_df) + 1
            last_col_res = len(resumen_df.columns)
            tabla_res = Table(displayName="TablaResumenExtras", ref=f"A1:{chr(64+last_col_res)}{last_row_res}")
            tabla_res.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False)
            hoja_resumen.add_table(tabla_res)
            for col in hoja_resumen.columns:
                hoja_resumen.column_dimensions[col[0].column_letter].width = \
                    max((len(str(c.value or "")) for c in col), default=0) + 3
        else:
            # Si descarga un empleado específico, mantener el comportamiento actual
            df_total.to_excel(writer, index=False, sheet_name="Horas Extras")
            resumen_df = pd.DataFrame(resumen).drop(columns=["extra_min"])
            resumen_df.to_excel(writer, index=False, sheet_name="Resumen")

            hoja = writer.sheets["Horas Extras"]
            last_row = len(df_total) + 1
            last_col = len(df_total.columns)
            tabla = Table(displayName="TablaExtras", ref=f"A1:{get_column_letter(last_col)}{last_row}")
            tabla.tableStyleInfo = TableStyleInfo(name="TableStyleLight11",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=False, showColumnStripes=False)
            hoja.add_table(tabla)
            for col in hoja.columns:
                hoja.column_dimensions[col[0].column_letter].width = \
                    max((len(str(c.value or "")) for c in col), default=0) + 3
            for i in range(1, last_row + 1):
                hoja.row_dimensions[i].height = 22
            fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for cell in hoja[last_row]:
                cell.fill = fill; cell.font = Font(bold=True)

            hoja_resumen = writer.sheets["Resumen"]
            last_row_res = len(resumen_df) + 1
            last_col_res = len(resumen_df.columns)
            tabla_res = Table(displayName="TablaResumenExtras", ref=f"A1:{chr(64+last_col_res)}{last_row_res}")
            tabla_res.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False)
            hoja_resumen.add_table(tabla_res)
            for col in hoja_resumen.columns:
                hoja_resumen.column_dimensions[col[0].column_letter].width = \
                    max((len(str(c.value or "")) for c in col), default=0) + 3
    output.seek(0)
    return send_file(output, download_name="horas_extras.xlsx", as_attachment=True)


@app.route("/descargar_llegadas")
def descargar_llegadas():
    nombre = request.args.get("nombre", "todos")
    df     = MEMORY.get("df")
    if df is None or df.empty:
        return "No hay datos cargados", 400

    df = df[~df["Nombre"].str.contains("TOTAL HORAS EXTRAS", na=False)].copy()
    if nombre != "todos":
        df = df[df["Nombre"] == nombre].copy()

    columnas = ["Nombre", "Fecha", "Día", "Entrada", "Sal. Almuerzo", "Reg. Almuerzo", "T. Almuerzo"]
    df = df[[c for c in columnas if c in df.columns]]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Llegadas")
        hoja     = writer.sheets["Llegadas"]
        last_row = len(df) + 1
        last_col = len(df.columns)
        tabla    = Table(displayName="TablaLlegadas", ref=f"A1:{get_column_letter(last_col)}{last_row}")
        tabla.tableStyleInfo = TableStyleInfo(name="TableStyleLight11",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False)
        hoja.add_table(tabla)
        for col in hoja.columns:
            hoja.column_dimensions[col[0].column_letter].width = \
                max((len(str(c.value or "")) for c in col), default=0) + 3
        for i in range(1, last_row + 1):
            hoja.row_dimensions[i].height = 22
    output.seek(0)
    return send_file(output, download_name=f"llegadas_{nombre}.xlsx", as_attachment=True)


@app.route("/descargar_almuerzo")
def descargar_almuerzo():
    nombre = request.args.get("nombre", "todos")
    df     = MEMORY.get("df")
    if df is None or df.empty:
        return "No hay datos cargados", 400

    df = df[~df["Nombre"].str.contains("TOTAL HORAS EXTRAS", na=False)].copy()
    if nombre != "todos":
        df = df[df["Nombre"] == nombre].copy()

    output = construir_almuerzo_excel(df)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name=f"almuerzo_{nombre}.xlsx",
        as_attachment=True,
    )


# ─────────────────────────────────────────────────────────────
# AJUSTES / LIMPIAR
# ─────────────────────────────────────────────────────────────

@app.route("/ajustes")
def ajustes():
    return render_template("ajustes.html", horarios=HORARIOS_SEDES,
                           permisos=len(MEMORY.get("permisos", [])),
                           sedes=list(HORARIOS_SEDES.keys()))

@app.route("/limpiar")
def limpiar():
    MEMORY["df"]      = None
    MEMORY["permisos"] = []
    return redirect(url_for("index"))


# ─────────────────────────────────────────────────────────────
# CONFIG (login, dashboard, sedes, horarios)
# ─────────────────────────────────────────────────────────────

@app.route("/config/login", methods=["GET", "POST"])
def config_login():
    if request.method == "POST":
        if request.form.get("password") == CONFIG_PASSWORD:
            session["config_auth"] = True
            return redirect(url_for("configuraciones"))
        return render_template("config_login.html", error="Contraseña incorrecta")
    return render_template("config_login.html")

@app.route("/config/logout")
def config_logout():
    session.pop("config_auth", None)
    return redirect(url_for("index"))

@app.route("/configuraciones", methods=["GET", "POST"])
@require_config_auth
def configuraciones():
    mensaje_config = None
    if request.method == "POST":
        umbral = request.form.get("umbral_extras_minutos", "")
        try:
            umbral_int = int(umbral)
            if umbral_int < 0:
                umbral_int = 0
            SETTINGS["umbral_extras_minutos"] = umbral_int
            guardar_settings(SETTINGS)
            mensaje_config = f"Umbral de horas extras actualizado a {umbral_int} minutos."
        except ValueError:
            mensaje_config = "Ingrese un valor entero válido para el umbral."

    empleados = cargar_empleados()
    sedes_list = list(HORARIOS_SEDES.keys())
    return render_template("configuraciones.html",
                           sedes=sedes_list, horarios=HORARIOS_SEDES,
                           empleados=empleados, total_sedes=len(sedes_list),
                           total_empleados=len(empleados),
                           permisos=MEMORY.get("permisos", []),
                           settings=SETTINGS, mensaje_config=mensaje_config)

@app.route("/config/dashboard")
@require_config_auth
def config_dashboard():
    permisos   = MEMORY.get("permisos", [])
    sedes_list = list(HORARIOS_SEDES.keys())
    empleados  = cargar_empleados()
    df         = MEMORY.get("df")
    empleados_totales = len(empleados)

    if df is None or df.empty:
        resumen = {
            "registros": 0, "empleados_presentes": 0,
            "tardanzas": 0, "horas_extras": 0,
            "top_tardanzas": [], "top_extras": [],
            "chart_asistencia": {"labels": ["Presentes", "Sin registro"], "data": [0, empleados_totales]},
            "chart_tardanzas":  {"labels": [], "data": []},
            "chart_extras":     {"labels": [], "data": []},
        }
    else:
        df_a = df[~df["Nombre"].str.contains("TOTAL HORAS EXTRAS", na=False)].copy()
        df_a["tardanza_min"] = df_a["Tardanza"].apply(parse_hhmm)
        df_a["extras_min"]   = df_a["Horas extras"].apply(parse_hhmm)

        presentes   = df_a[df_a["Entrada"] != "--:--"]["Nombre"].nunique()
        tardanzas   = df_a[df_a["tardanza_min"] > 0]
        extras      = df_a[df_a["extras_min"] > 0]
        tard_emp    = tardanzas.groupby("Nombre")["tardanza_min"].sum().sort_values(ascending=False).head(5)
        extras_emp  = extras.groupby("Nombre")["extras_min"].sum().sort_values(ascending=False).head(5)

        resumen = {
            "registros": len(df_a), "empleados_presentes": presentes,
            "tardanzas": tardanzas["Nombre"].nunique(),
            "horas_extras": extras["Nombre"].nunique(),
            "top_tardanzas": tardanzas.sort_values("tardanza_min", ascending=False)
                                      .head(5)[["Nombre","Fecha","Día","Tardanza"]].to_dict("records"),
            "top_extras":    extras.sort_values("extras_min", ascending=False)
                                   .head(5)[["Nombre","Fecha","Día","Horas extras"]].to_dict("records"),
            "chart_asistencia": {"labels": ["Presentes","Sin registro"],
                                 "data": [presentes, max(empleados_totales - presentes, 0)]},
            "chart_tardanzas":  {"labels": tard_emp.index.tolist(), "data": tard_emp.tolist()},
            "chart_extras":     {"labels": extras_emp.index.tolist(), "data": extras_emp.tolist()},
        }

    return render_template("config_dashboard.html",
                           total_permisos=len(permisos), total_sedes=len(sedes_list),
                           total_empleados=empleados_totales,
                           permisos=permisos, resumen=resumen)

@app.route("/config/sedes", methods=["GET", "POST"])
@require_config_auth
def config_sedes():
    if request.method == "POST":
        accion      = request.form.get("accion")
        nombre_sede = request.form.get("nombre_sede", "").strip().lower()
        if accion == "crear" and nombre_sede and nombre_sede not in HORARIOS_SEDES:
            HORARIOS_SEDES[nombre_sede] = {
                d: {"entrada": "08:00", "salida": "17:00", "descontar_almuerzo": True}
                for d in ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"]
            }
            guardar_config_sedes(HORARIOS_SEDES)
        elif accion == "eliminar" and nombre_sede in HORARIOS_SEDES and len(HORARIOS_SEDES) > 1:
            del HORARIOS_SEDES[nombre_sede]
            guardar_config_sedes(HORARIOS_SEDES)
    return render_template("config_sedes.html", sedes=list(HORARIOS_SEDES.keys()))

@app.route("/config/horarios/<sede>", methods=["GET", "POST"])
@require_config_auth
def config_horarios(sede):
    if sede not in HORARIOS_SEDES:
        return "Sede no encontrada", 404
    horario_sede = HORARIOS_SEDES[sede]
    if request.method == "POST":
        for dia in list(horario_sede.keys()):
            entrada    = request.form.get(f"{dia}_entrada", "")
            salida     = request.form.get(f"{dia}_salida", "")
            descontar  = request.form.get(f"{dia}_descontar") is not None
            if entrada and salida:
                horario_sede[dia] = {"entrada": entrada, "salida": salida,
                                     "descontar_almuerzo": descontar}
        guardar_config_sedes(HORARIOS_SEDES)
    return render_template("config_horarios.html", sede=sede, horario=horario_sede)


if __name__ == "__main__":
    app.run(debug=True)